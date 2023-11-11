from dataclasses import dataclass
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from rich.progress import track
import xlwings as xw



def find_footing_files(directory: Path, pattern: str) -> list:
    """
    Function for finding files in the big folder
    that have *footing in them
    """
    filepaths = []
    for filepath in Path(directory).rglob(pattern):
        # print(filepath)
        filepaths.append(filepath)
    return filepaths

def write_footings_to_excel(filepaths: list, workbook_path: Path, startdate: datetime, enddate: datetime):
    """
    takes the filepaths of all the footings and writes the names of the projects
    in the bad date range into a new workbook
    """
    towb = load_workbook(workbook_path)
    tows = towb.active
    i = 1
    for filepath in filepaths:
        #print(filepath)
        wb = load_workbook(filepath)
        ws = wb.active
        project_name = ws['C2'].value
        print(project_name)
        date = ws['C3'].value
        #date = datetime.strptime(date_str, '%d-%m-%Y')
        if date >= startdate:
            if date <= enddate:
                i = i+1
                tows[f'A{i}'].value = i-1
                tows[f'B{i}'].value = project_name
                tows[f'C{i}'].value = date
            else:
                print("not this one")
        else:
            print("not this one")
        
    towb.save(workbook_path)
    return 


def update_rebar(filename: Path, fy_new: int) -> list[dict]:
    """
    This function opens a spreadsheet, takes the rebar design,
    updates the yield strenght of the steel, gets the new rebar design,
    reports changes 
    """
    wb = xw.Book(filename) 
    ws = wb.sheets[0]
    
    original_rebar = {
        'nlong': ws['E67'].value,
        'nwidth': ws['E68'].value,
        'typelong': ws['E69'].value,
        'typewidth': ws['E70'].value
    }
    #print(f"Original values: {original_rebar}")
    
    ws['E11'].value = fy_new   

    updated_rebar = {
        'nlong': ws['E67'].value,
        'nwidth': ws['E68'].value,
        'typelong': ws['E69'].value,
        'typewidth': ws['E70'].value
    }
    #print(f"Updated values: {updated_rebar}")
    
    changes = {key: updated_rebar[key] != original_rebar[key] for key in original_rebar}
    #print(f"Changes: {changes}")
    wb.close()
    return [original_rebar, updated_rebar, changes]


def filter_footings(filepaths: list, workbook_path_nochanges: Path, workbook_path_changes: Path, startdate: datetime, enddate: datetime, fynew: int):
    """
    takes the filepaths of all the footings and writes the names of the projects
    in the bad date range into a new workbook
    """
    towb_nch = load_workbook(workbook_path_nochanges)
    tows_nch = towb_nch.active
    towb_ch = load_workbook(workbook_path_changes)
    tows_ch = towb_ch.active
    i = 1
    idx = 2
    for filepath in filepaths:
        #print(filepath)
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            project_name = ws['C2'].value
            print(project_name)
            date = ws['C3'].value
            if date >= startdate:
                if date <= enddate:
                    results = update_rebar(filepath, fynew)
                    #print(results)
                    changes = results[2]
                    #print(changes)
                    if any(value for value in changes.values()):
                        #write values to the sheet with changes
                        idx = idx+1
                        tows_ch[f'A{idx}'].value = i-2
                        tows_ch[f'B{idx}'].value = project_name
                        tows_ch[f'C{idx}'].value = date
                        tows_ch[f'D{idx}'].value = results[0]["nlong"]
                        tows_ch[f'E{idx}'].value = results[0]["nwidth"]
                        tows_ch[f'F{idx}'].value = results[0]["typelong"]
                        tows_ch[f'G{idx}'].value = results[0]["typewidth"]
                        tows_ch[f'H{idx}'].value = results[1]["nlong"]
                        tows_ch[f'I{idx}'].value = results[1]["nwidth"]
                        tows_ch[f'J{idx}'].value = results[1]["typelong"]
                        tows_ch[f'K{idx}'].value = results[1]["typewidth"]
                        tows_ch[f'L{idx}'].value = changes["nlong"]
                        tows_ch[f'M{idx}'].value = changes["nwidth"]
                        tows_ch[f'N{idx}'].value = changes["typelong"]
                        tows_ch[f'O{idx}'].value = changes["typewidth"]
                        #save sheet with new values under a new name
                        original_stem = filepath.stem
                        updated_stem = f"{original_stem}_updated"
                        new_path = filepath.with_name(f"{updated_stem}.xlsx")
                        wb.save(new_path)
                        wb.close()
                        
                    else:
                        #write values to the sheet with no changes
                        i = i+1
                        tows_nch[f'A{i}'].value = i-1
                        tows_nch[f'B{i}'].value = project_name
                        tows_nch[f'C{i}'].value = date
                        wb.close()
                else:
                    i = i+1
                    tows_nch[f'A{i}'].value = i-1
                    tows_nch[f'B{i}'].value = project_name
                    tows_nch[f'C{i}'].value = date
                    wb.close()

            else:
                i = i+1
                tows_nch[f'A{i}'].value = i-1
                tows_nch[f'B{i}'].value = project_name
                tows_nch[f'C{i}'].value = date
                wb.close()
                
        finally:
            wb.close()

    towb_ch.save(workbook_path_changes)
    towb_nch.save(workbook_path_nochanges)
    return 

